import os
import pathlib
import datetime
import time

import openpyxl
import json
import openpyxl.styles

import traceback

from PyQt5.QtCore import QThread, pyqtSignal


class MyException(Exception):
    pass


class GetReport(QThread):  # Если требуется вставить колонтитулы
    progress = pyqtSignal(int)  # Сигнал для прогресс бара
    status = pyqtSignal(str)  # Сигнал для статус бара
    messageChanged = pyqtSignal(str, str)
    errors = pyqtSignal()

    def __init__(self, incoming_data):  # Список переданных элементов.
        QThread.__init__(self)
        self.queue = incoming_data['queue']
        self.logging = incoming_data['logging']

    def run(self):
        try:
            self.logging.info("\n***********************************Report***********************************\n")
            logs_files = {os.path.getctime(pathlib.Path('logs', file)): file
                          for file in os.listdir(pathlib.Path('logs'))}
            logs_files = dict(sorted(logs_files.items(), reverse=True))
            if len(logs_files) == 0 or (len(logs_files) == 1 and
                                        os.path.getsize(pathlib.Path('logs',
                                                                     logs_files[list(logs_files.keys())[0]])) == 0):
                raise MyException('Нет файлов для формирования отчёта, логи удалены или программа ни разу не работала')
            data_for_report = ''
            self.logging.info('Считываем файл лога для получения информации')
            for file in logs_files:
                for_report = False
                with open(pathlib.Path('logs', logs_files[file]), 'r') as f:
                    lines = f.readlines()
                for line in reversed(lines):
                    if '*****Finish*****' in line:
                        for_report = True
                    if for_report and 'for_report:' in line:
                        data_for_report = line
                        break
                    if '*****Start*****' in line and for_report:
                        break
                if len(data_for_report):
                    break
                print(os.path.getsize(pathlib.Path('logs', logs_files[file])))
                if for_report:
                    break
            if len(data_for_report):
                pass
            else:
                raise MyException('В текущих логах не найдено успешного выполнения программы, невозможно сделать отчёт')
            self.logging.info(data_for_report)
            self.logging.info('Формируем отчёт')
            report_data = json.loads(data_for_report.partition('for_report:')[2])
            thin = openpyxl.styles.Side(border_style="thin", color="000000")
            medium = openpyxl.styles.Side(border_style="medium", color="000000")
            name_temp_file = str(int(time.time())) + '.xlsx'
            wb = openpyxl.Workbook()
            ws = wb.active
            for i in range(1, 5):
                for j in range(1, 13):
                    if j == 8 and i == 1:
                        ws.cell(i, j).border = openpyxl.styles.Border(top=thin, left=thin, right=medium, bottom=thin)
                    elif i == 2 and j == 8:
                        ws.cell(i, j).border = openpyxl.styles.Border(top=thin, left=thin, right=medium, bottom=medium)
                    elif i == 4 and j == 12:
                        ws.cell(i, j).border = openpyxl.styles.Border(top=medium, left=thin, right=medium, bottom=medium)
                    elif i == 2:
                        ws.cell(i, j).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=medium)
                    elif i == 3:
                        ws.cell(i, j).border = openpyxl.styles.Border(bottom=medium)
                    elif i == 4:
                        ws.cell(i, j).border = openpyxl.styles.Border(top=medium, left=thin, right=thin, bottom=medium)
                    else:
                        ws.cell(i, j).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(i, j).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center",
                                                                        wrap_text=True)
            for j in range(65, 79):
                ws.column_dimensions[chr(j)].width = 37 if j == 65 else 17
            ws.cell(1, 1).value = 'Дата формирования отчёта:'
            ws.cell(1, 2).value = 'Время начала работы программы:'
            ws.cell(1, 3).value = 'Время выполнения программы:'
            ws.cell(1, 4).value = 'Всего документов:'
            ws.cell(1, 5).value = 'Обработано успешно:'
            ws.cell(1, 6).value = 'Пропущено (не было в выгрузке):'
            ws.cell(1, 7).value = 'Обработано с ошибкой:'
            ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
            ws.cell(1, 8).value = 'Ошибки в выгрузке:'
            ws.cell(4, 1).value = 'Имя документа'
            ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
            ws.cell(4, 2).value = 'Ошибки при заполнении'
            ws.cell(4, 6).value = 'Серийный номер'
            ws.cell(4, 7).value = 'Состав'
            ws.cell(4, 8).value = 'Изготовитель'
            ws.cell(4, 9).value = 'Модель'
            ws.cell(4, 10).value = 'Заводской номер'
            ws.cell(4, 11).value = 'Копирование'
            ws.cell(4, 12).value = 'Удаление'
            index_device = 5
            time_start = datetime.datetime.strptime(report_data['info']['time_start'].partition(' ')[0],
                                                    '%Y-%m-%d').date()
            self.logging.info('Бежим по ключам')
            for key in report_data:
                if key == 'info':
                    ws.cell(2, 1).value = time_start
                    ws.cell(2, 2).value = \
                        datetime.datetime.strptime(report_data[key]['time_start'].partition(' ')[2].rpartition('.')[0],
                                                   '%H:%M:%S').time()
                    ws.cell(2, 3).value = \
                        datetime.datetime.strptime(report_data[key]['work_time'].rpartition('.')[0], '%H:%M:%S').time()
                    ws.cell(2, 4).value = report_data[key]['all_file']
                    ws.cell(2, 5).value = report_data[key]['success']
                    ws.cell(2, 6).value = report_data[key]['pass']
                    ws.cell(2, 7).value = report_data[key]['errors']
                elif key == 'unloading':
                    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=12)
                    if report_data[key]['errors']:
                        ws.cell(2, 8).fill = openpyxl.styles.PatternFill(start_color='FFFF00',
                                                                         end_color='FFFF00',
                                                                         fill_type="solid")
                        if len(report_data[key]['text_err']) > 1:
                            ws.row_dimensions[2].height = 14.4 * len(report_data[key]['text_err'])
                        ws.cell(2, 8).value = '\n'.join(report_data[key]['text_err'])
                    else:
                        ws.cell(2, 8).value = 'Нет ошибок'
                else:
                    for j in range(1, 13):
                        if index_device == len(report_data) + 2 and j == 12:
                            ws.cell(index_device, j).border = openpyxl.styles.Border(top=thin, left=thin,
                                                                                     right=medium, bottom=medium)
                        elif index_device == len(report_data) + 2:
                            ws.cell(index_device, j).border = openpyxl.styles.Border(top=thin, left=thin,
                                                                                     right=thin, bottom=medium)
                        elif j == 12:
                            ws.cell(index_device, j).border = openpyxl.styles.Border(top=thin, left=thin,
                                                                                     right=medium, bottom=thin)
                        else:
                            ws.cell(index_device, j).border = openpyxl.styles.Border(top=thin,
                                                                                     left=thin, right=thin, bottom=thin)
                        ws.cell(index_device, j).alignment = openpyxl.styles.Alignment(horizontal="center",
                                                                                       vertical="center",
                                                                                       wrap_text=True)
                    ws.cell(index_device, 1).value = key
                    ws.merge_cells(start_row=index_device, start_column=2, end_row=index_device, end_column=5)
                    if report_data[key]['errors']:
                        for j in range(1, 13):
                            ws.cell(index_device, j).fill = openpyxl.styles.PatternFill(start_color='FFFF00',
                                                                                        end_color='FFFF00',
                                                                                        fill_type="solid")
                            # ws.cell(index_device, j).font = openpyxl.styles.Font(color='FFDF00')
                        if len(report_data[key]['text_err']) > 1:
                            ws.row_dimensions[index_device].height = 14.4 * len(report_data[key]['text_err'])
                        ws.cell(index_device, 2).value = '\n'.join(report_data[key]['text_err'])
                    else:
                        ws.cell(index_device, 2).value = 'Нет ошибок'
                    ws.cell(index_device, 6).value = 'Заполнено' if report_data[key]['s/n'] else 'Не заполнено'
                    ws.cell(index_device, 7).value = 'Заполнено' if report_data[key]['compound'] else 'Не заполнено'
                    ws.cell(index_device, 8).value = 'Заполнено' if report_data[key]['manufacturer'] else 'Не заполнено'
                    ws.cell(index_device, 9).value = 'Заполнено' if report_data[key]['model'] else 'Не заполнено'
                    ws.cell(index_device, 10).value = 'Заполнено'\
                        if report_data[key]['factory number'] else 'Не заполнено'
                    ws.cell(index_device, 11).value = 'Скопирован' if report_data[key]['copy'] else 'Не скопирован'
                    ws.cell(index_device, 12).value = 'Удалён' if report_data[key]['del'] else 'Не удалён'
                    index_device += 1
            wb.save(pathlib.Path(pathlib.Path.cwd(), 'report', name_temp_file))
            os.startfile(pathlib.Path(pathlib.Path.cwd(), 'report', name_temp_file))
            self.status.emit('Отчёт готов!')  # Посылаем значние если готово
            self.logging.info("\n***********************************End report***********************************\n")
        except MyException as e:
            self.status.emit('Ошибка формирования отчёта.')  # Посылаем значние если ошибка
            self.logging.error("Ошибка:\n " + str(e) + '\n' + traceback.format_exc())
            self.messageChanged.emit('УПС!', str(e))
            self.logging.info("\n***********************************Error report***********************************\n")
        except BaseException as e:
            self.status.emit('Не удается открыть отчёт')  # Сообщение в статус бар
            self.logging.error("Ошибка:\n " + str(e) + '\n' + traceback.format_exc())
            self.logging.info("\n***********************************Error report***********************************\n")
